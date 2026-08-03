"""extract_envelope_surfaces — surfaces d'enveloppe depuis la géométrie IFC.

Calcule, pour alimenter l'annexe I3F « Extraction surface enveloppe » :

- **Superficie des façades** : aire verticale des murs extérieurs + murs-rideaux
  (parement extérieur, ouvertures déduites) + les menuiseries (qui occupent le
  plan de façade) → aire de façade *brute*.
- **Superficie des menuiseries** : aire des IfcWindow + IfcDoor extérieures
  (largeur × hauteur d'emprise).
- **SHAB** : surface habitable = somme des surfaces nettes des pièces, hors
  annexes (cave, cellier, parking, local technique, extérieur).
- **ratio FAC/SHAB** : compacité de l'enveloppe. Définition **unique** dans tous
  les modes — ``superficie_facades_nette_m2 / shab_m2``, menuiseries **exclues**,
  celle que le livrable Excel calcule. Deux définitions concurrentes du même
  indicateur ont circulé (0,92 dans le classeur, 1,05 dans le contrat) : c'est
  précisément ce que cette règle unique interdit.

Produit un dict + un classeur .xlsx au format lu par
``audit_bim.reporting.avp_sources.read_enveloppe`` (libellé en colonne A,
valeur numérique en colonne B ; table détaillée ancrée sur « Composant »).
"""

from __future__ import annotations

import re
import warnings

import ifcopenshell.util.element as ue
from bim_core.contracts import SCHEMA_ENVELOPE_QUANTITIES_V1

from .. import ifc_utils
from ..contracts import contract_source, utc_now_iso
from .spaces import normalize_room_type

_WALL_TYPES = ("IfcWall", "IfcWallStandardCase")
_CURTAIN_TYPES = ("IfcCurtainWall",)
# Pièces exclues de la surface habitable (annexes non chauffées / extérieur).
_SHAB_EXCLUDE = {"cave", "cellier", "parking", "technique", "exterieur"}
# Annexes non habitables exclues de la SHAB **I3F** (mode calque) — jeu de
# l'extraction de référence Tarare 0546L : cellier, cave, balcon, garage,
# escalier, local technique.
#
# Ces libellés métier ne sont PAS les valeurs rendues par
# ``normalize_room_type`` : « balcon » y devient ``exterieur``, « local
# technique » ``technique``, et « garage » / « escalier » n'ont aucun motif et
# retombent sur ``autre``. Comparer les libellés métier au type normalisé
# n'exclurait donc que ``cave`` et ``cellier``, et laisserait un garage zoné
# gonfler la SHAB — donc fausser le ratio FAC/SHAB. D'où deux niveaux.
_SHAB_EXCLUDE_I3F_TYPES = frozenset({"cave", "cellier", "exterieur", "technique"})
# Repli sur le libellé brut, pour ce que la normalisation ne distingue pas.
_SHAB_EXCLUDE_I3F_RAW = re.compile(r"garage|escalier", re.I)


def _is_i3f_shab_excluded(label: str | None) -> bool:
    """La pièce est-elle une annexe non habitable au sens I3F ?"""
    if normalize_room_type(label) in _SHAB_EXCLUDE_I3F_TYPES:
        return True
    return bool(_SHAB_EXCLUDE_I3F_RAW.search(label or ""))


_WALL_LIKE = {"IfcWall", "IfcWallStandardCase", "IfcCurtainWall"}


def _external_wall_guids(model) -> tuple[set[str], str]:
    """GlobalId des murs extérieurs.

    Méthode primaire : ``IfcRelSpaceBoundary`` marquées ``EXTERNAL`` (fiable même
    quand ``Pset_WallCommon.IsExternal`` est absent/faux — cas fréquent des
    exports ArchiCAD). Repli : flag ``IsExternal=True``.
    """
    guids: set[str] = set()
    for rel in model.by_type("IfcRelSpaceBoundary"):
        if getattr(rel, "InternalOrExternalBoundary", None) != "EXTERNAL":
            continue
        el = getattr(rel, "RelatedBuildingElement", None)
        if el is not None and el.is_a() in _WALL_LIKE:
            guids.add(el.GlobalId)
    method = "space_boundaries"
    if not guids:  # repli sur le flag IsExternal
        method = "is_external_flag"
        for t in (*_WALL_TYPES, *_CURTAIN_TYPES):
            for el in model.by_type(t):
                if ifc_utils.is_external(el) is True:
                    guids.add(el.GlobalId)
    return guids, method


def _wall_type(el) -> str:
    """Type **métier** d'un mur.

    Ordre de résolution, du plus signifiant au moins signifiant :

    1. le **type IFC** (``IfcWallType.Name``), atteint via
       ``IfcRelDefinesByType``. C'est là que vivent les noms de composants
       ArchiCAD (« ME_R+1_Enduit_recoupement 530 x 2850 ») comme les types
       Revit (« Mur de base:BARDAGE BOIS 20mm + VENTIL 50mm »).
       ``ifcopenshell.util.element.get_type`` est utilisé plutôt que
       ``IsTypedBy`` : cette relation n'existe qu'en IFC4, alors que les
       maquettes I3F — ArchiCAD comme Revit — sont en **IFC2X3**
       (``IsDefinedBy`` → ``IfcRelDefinesByType``) ;
    2. ``ObjectType`` — le repli des exports Revit qui ne portent pas de
       ``IfcWallType`` nommé ;
    3. ``PredefinedType``, en **dernier recours** : sur ArchiCAD il vaut
       ``ELEMENTEDWALL`` pour tous les murs et écraserait la décomposition
       métier en un type unique.

    Le ``Name`` de l'**instance** reste en ultime filet, après
    ``PredefinedType`` : sur Revit il porte l'identifiant de l'élément
    (« Mur de base:MUR ENDUIT 20 mm:3566323 ») et produirait un type distinct
    par mur — inexploitable comme clé de regroupement.
    """
    t = ue.get_type(el)
    tn = getattr(t, "Name", None) if t is not None else None
    if isinstance(tn, str) and tn.strip():
        return tn.strip()
    v = getattr(el, "ObjectType", None)
    if isinstance(v, str) and v.strip():
        return v.strip()
    pt = getattr(el, "PredefinedType", None)
    if isinstance(pt, str) and pt.strip() and pt != "NOTDEFINED":
        return pt.strip()
    v = getattr(el, "Name", None)
    if isinstance(v, str) and v.strip():
        return v.strip()
    return el.is_a()


def _wall_layer(el) -> str | None:
    """Calque ArchiCAD d'un mur.

    Source primaire : pset ``ArchiCADProperties`` (``Calque`` ou ``Layer``).
    Repli : ``IfcPresentationLayerAssignment`` porté par la représentation.
    """
    ac = (ue.get_psets(el) or {}).get("ArchiCADProperties") or {}
    for key in ("Calque", "Layer"):
        v = ac.get(key)
        if isinstance(v, str) and v.strip():
            return v.strip()
    rep = getattr(el, "Representation", None)
    for r in getattr(rep, "Representations", None) or []:
        for la in getattr(r, "LayerAssignments", None) or []:
            name = getattr(la, "Name", None)
            if isinstance(name, str) and name.strip():
                return name.strip()
    return None


def _wall_side_area(el) -> float | None:
    """NetSideArea (Qto) sinon aire de parement géométrique."""
    a = ifc_utils.quantity(el, "NetSideArea", "GrossSideArea", "GrossArea")
    if a is None:
        a = ifc_utils.vertical_face_area(el)
    return a


def _finalize_by_type(buckets: dict) -> list[dict]:
    """Lignes ``par_type`` aux **noms canoniques du contrat V1**.

    Les anciens noms (``netsidearea_m2``, ``nombre``, ``etages`` concaténés en
    une chaîne) étaient précisément les alias que bim-core doit normaliser à la
    lecture d'un fichier historique : un producteur à jour ne les émet plus.
    """
    return sorted(
        (
            {
                "type": b["type"],
                "etages": sorted(b["etages"]),
                "net_side_area_m2": round(b["netsidearea_m2"], 2),
                "n": b["nombre"],
            }
            for b in buckets.values()
        ),
        key=lambda x: x["type"],
    )


def _facades(model) -> dict:
    """Décomposition des murs par type.

    - ``par_type`` : murs **extérieurs** (façades), agrégés par type — c'est le
      total métier de l'annexe MOA (colonne D « Archicad BQ NetSideArea »).
    - ``hors_filtre_type`` : murs **non retenus** par le filtre extérieur — exclus
      du total métier, exposés à titre diagnostic (jamais dans le total façade).
    - ``superficie_calque_total_m2`` : NetSideArea de **tous** les murs (retenus +
      hors filtre) — le total « brut » avant filtrage.
    """
    ext_guids, method = _external_wall_guids(model)
    par: dict[str, dict] = {}
    hors: dict[str, dict] = {}
    facade_net = calque_total = 0.0
    n_ext = n_geom_fallback = 0

    for t in (*_WALL_TYPES, *_CURTAIN_TYPES):
        for el in model.by_type(t):
            a = _wall_side_area(el)
            if not a:
                continue
            if (
                ifc_utils.quantity(el, "NetSideArea", "GrossSideArea", "GrossArea")
                is None
            ):
                n_geom_fallback += 1
            calque_total += a
            is_ext = el.GlobalId in ext_guids
            bucket = par if is_ext else hors
            wt = _wall_type(el)
            b = bucket.setdefault(
                wt, {"type": wt, "etages": set(), "netsidearea_m2": 0.0, "nombre": 0}
            )
            b["netsidearea_m2"] += a
            b["nombre"] += 1
            st = ifc_utils.storey_name(el)
            if st:
                b["etages"].add(str(st))
            if is_ext:
                facade_net += a
                n_ext += 1

    return {
        "par_type": _finalize_by_type(par),
        "hors_filtre_type": _finalize_by_type(hors),
        "facade_net": round(facade_net, 2),
        "calque_total": round(calque_total, 2),
        "method": method,
        "n_ext": n_ext,
        "n_geom_fallback": n_geom_fallback,
    }


def _facades_by_geometric_type(model, type_re) -> dict:
    """Décomposition **Revit** : murs extérieurs géométriques, filtrés par type.

    Le pendant de :func:`_facades_by_layer` pour les maquettes **sans calque**.
    Un export Revit n'expose pas de calque ArchiCAD : la sélection I3F
    « calque + type » n'y retient rien, et la sélection purement géométrique y
    retient trop — chaque façade est modélisée en **murs superposés** (structure,
    isolant, peau extérieure), si bien que sommer les murs extérieurs compte la
    même façade trois ou quatre fois.

    ``type_pattern`` tranche cette superposition : il désigne la couche qui
    représente la façade. Les autres types restent listés en
    ``hors_filtre_type``, **hors du total métier** — visibles pour l'audit, mais
    jamais additionnés.
    """
    ext_guids, method_ext = _external_wall_guids(model)
    par: dict[str, dict] = {}
    hors: dict[str, dict] = {}
    wall_types: dict[int, str] = {}
    facade_net = calque_total = 0.0
    n_retenus = n_ext = n_geom_fallback = 0

    for t in (*_WALL_TYPES, *_CURTAIN_TYPES):
        for el in model.by_type(t):
            if el.GlobalId not in ext_guids:
                continue
            n_ext += 1
            a = _wall_side_area(el)
            if not a:
                continue
            if (
                ifc_utils.quantity(el, "NetSideArea", "GrossSideArea", "GrossArea")
                is None
            ):
                n_geom_fallback += 1
            wt = _wall_type(el)
            wall_types[el.id()] = wt
            retenu = type_re is None or bool(type_re.search(wt))
            b = (par if retenu else hors).setdefault(
                wt, {"type": wt, "etages": set(), "netsidearea_m2": 0.0, "nombre": 0}
            )
            b["netsidearea_m2"] += a
            b["nombre"] += 1
            st = ifc_utils.storey_name(el)
            if st:
                b["etages"].add(str(st))
            calque_total += a
            if retenu:
                facade_net += a
                n_retenus += 1

    return {
        "par_type": _finalize_by_type(par),
        "hors_filtre_type": _finalize_by_type(hors),
        "facade_net": round(facade_net, 2),
        # Total des murs extérieurs AVANT filtre de type : c'est la valeur qui
        # compte plusieurs fois la même façade. Diagnostic, jamais un résultat.
        "calque_total": round(calque_total, 2),
        "method": "geometric_type_filter",
        "method_exterieur": method_ext,
        "n_ext": n_retenus,
        "n_murs_exterieurs": n_ext,
        "n_geom_fallback": n_geom_fallback,
        "wall_types": wall_types,
    }


def _facades_by_layer(model, layer_re, type_re) -> dict:
    """Décomposition I3F : murs sélectionnés par **calque**, puis filtrés par
    **nom de type** métier.

    Reproduit l'extraction de référence Tarare 0546L :

    - le calque (ex. « 221 - MURS - Extérieurs périphériques ») délimite les
      murs d'enveloppe — un filtre géométrique « extérieur » ne le remplace pas,
      il retient aussi habillages et ouvrages annexes ;
    - au sein du calque, ``type_pattern`` (ex. ``^ME[ _]``) distingue les murs
      extérieurs des habillages (zinc, alu, bois, couvertines) : ces derniers
      restent listés en ``hors_filtre_type``, **hors du total métier** ;
    - le total façade est la somme des ``NetSideArea`` des types retenus —
      **sans** les menuiseries, qui ont leur propre total.
    """
    par: dict[str, dict] = {}
    hors: dict[str, dict] = {}
    wall_types: dict[int, str] = {}
    facade_net = calque_total = 0.0
    n_retenus = n_calque = n_sans_calque = n_geom_fallback = 0

    for t in (*_WALL_TYPES, *_CURTAIN_TYPES):
        for el in model.by_type(t):
            layer = _wall_layer(el)
            if layer is None:
                n_sans_calque += 1
                continue
            if not layer_re.search(layer):
                continue
            n_calque += 1
            a = _wall_side_area(el)
            if not a:
                continue
            if (
                ifc_utils.quantity(el, "NetSideArea", "GrossSideArea", "GrossArea")
                is None
            ):
                n_geom_fallback += 1
            wt = _wall_type(el)
            wall_types[el.id()] = wt
            retenu = type_re is None or bool(type_re.search(wt))
            b = (par if retenu else hors).setdefault(
                wt, {"type": wt, "etages": set(), "netsidearea_m2": 0.0, "nombre": 0}
            )
            b["netsidearea_m2"] += a
            b["nombre"] += 1
            st = ifc_utils.storey_name(el)
            if st:
                b["etages"].add(str(st))
            calque_total += a
            if retenu:
                facade_net += a
                n_retenus += 1

    return {
        "par_type": _finalize_by_type(par),
        "hors_filtre_type": _finalize_by_type(hors),
        "facade_net": round(facade_net, 2),
        "calque_total": round(calque_total, 2),
        "method": "layer_type_filter",
        "n_ext": n_retenus,
        "n_geom_fallback": n_geom_fallback,
        "n_murs_calque": n_calque,
        "n_murs_sans_calque": n_sans_calque,
        "wall_types": wall_types,
    }


def _menuiseries_of_walls(model, wall_types: dict[int, str]) -> dict:
    """Menuiseries **portées par les murs sélectionnés** (percement → remplissage).

    Contrairement au comptage global de toutes les baies extérieures, on ne
    retient ici que les fenêtres/portes qui remplissent une ouverture d'un mur
    du calque : c'est la définition de l'extraction I3F.
    """
    fills = {
        rel.RelatingOpeningElement.id(): rel.RelatedBuildingElement
        for rel in model.by_type("IfcRelFillsElement")
    }
    surf_fenetres = surf_portes = 0.0
    n = 0
    detail: list[dict] = []
    par_type_openings: dict[str, float] = {}
    for rel in model.by_type("IfcRelVoidsElement"):
        host = rel.RelatingBuildingElement
        if host is None or host.id() not in wall_types:
            continue
        fill = fills.get(rel.RelatedOpeningElement.id())
        if fill is None or not (fill.is_a("IfcWindow") or fill.is_a("IfcDoor")):
            continue
        res = _menuiserie_area(fill)
        if not res:
            continue
        w, h, area = res
        if fill.is_a("IfcWindow"):
            surf_fenetres += area
        else:
            surf_portes += area
        n += 1
        par_type_openings[wall_types[host.id()]] = (
            par_type_openings.get(wall_types[host.id()], 0.0) + area
        )
        detail.append(
            {
                "type": fill.is_a(),
                "name": fill.Name,
                "largeur_m": round(w, 2) if w else None,
                "hauteur_m": round(h, 2) if h else None,
                "surface_m2": round(area, 2),
            }
        )
    return {
        "total": surf_fenetres + surf_portes,
        "fenetres": surf_fenetres,
        "portes": surf_portes,
        "n": n,
        "detail": detail,
        "par_type": par_type_openings,
    }


def _shab_zoned(model) -> tuple[float, int, float]:
    """SHAB I3F : pièces **rattachées à une zone**, hors annexes non habitables.

    La double condition (zone + type de pièce) est ce qui distingue la SHAB de
    la simple somme des surfaces de pièces : une pièce hors zone n'appartient
    pas à un logement.
    """
    space_to_zones, _ = ifc_utils.zone_map(model)
    total = excluded = 0.0
    n = 0
    for sp in model.by_type("IfcSpace"):
        if not space_to_zones.get(sp.GlobalId):
            continue
        area = ifc_utils.quantity(sp, "NetFloorArea", "GrossFloorArea", "NetArea")
        if not area:
            continue
        label = ifc_utils.space_long_name(sp) or sp.Name or ""
        if _is_i3f_shab_excluded(label):
            excluded += area
            continue
        total += area
        n += 1
    return total, n, excluded


def _menuiserie_area(el) -> tuple[float, float, float] | None:
    """(largeur, hauteur, surface) d'une baie via attributs/Qto, sinon géométrie."""
    w = getattr(el, "OverallWidth", None)
    h = getattr(el, "OverallHeight", None)
    if isinstance(w, (int, float)) and isinstance(h, (int, float)) and w and h:
        return float(w), float(h), float(w) * float(h)
    a = ifc_utils.quantity(el, "Area")
    if a is not None:
        return None, None, a
    wh = ifc_utils.bbox_width_height(el)  # fallback géométrique
    if wh:
        return wh[0], wh[1], wh[0] * wh[1]
    return None


def _menuiseries(model) -> dict:
    """Aire des baies : fenêtres + portes extérieures, **splitée** par catégorie
    (colonnes MOA G « Surface des Fenêtres » / H « Surface des Portes »)."""
    surf_fenetres = surf_portes = 0.0
    n = 0
    detail: list[dict] = []
    for el in model.by_type("IfcWindow"):
        res = _menuiserie_area(el)
        if res:
            w, h, area = res
            surf_fenetres += area
            n += 1
            detail.append(
                {
                    "type": "IfcWindow",
                    "name": el.Name,
                    "largeur_m": round(w, 2) if w else None,
                    "hauteur_m": round(h, 2) if h else None,
                    "surface_m2": round(area, 2),
                }
            )
    for el in model.by_type("IfcDoor"):
        if ifc_utils.is_external(el) is True:
            res = _menuiserie_area(el)
            if res:
                w, h, area = res
                surf_portes += area
                n += 1
                detail.append(
                    {
                        "type": "IfcDoor(ext)",
                        "name": el.Name,
                        "largeur_m": round(w, 2) if w else None,
                        "hauteur_m": round(h, 2) if h else None,
                        "surface_m2": round(area, 2),
                    }
                )
    return {
        "total": surf_fenetres + surf_portes,
        "fenetres": surf_fenetres,
        "portes": surf_portes,
        "n": n,
        "detail": detail,
    }


def _shab(model) -> tuple[float, int]:
    """SHAB via NetFloorArea déclarée (Qto), hors annexes. Aucune géométrie."""
    total = 0.0
    n = 0
    for sp in model.by_type("IfcSpace"):
        label = getattr(sp, "LongName", None) or sp.Name or ""
        if normalize_room_type(label) in _SHAB_EXCLUDE:
            continue
        a = ifc_utils.quantity(sp, "NetFloorArea", "GrossFloorArea", "NetArea")
        if a:
            total += a
            n += 1
    return total, n


#: Modes de sélection des murs d'enveloppe. ``auto`` déduit le mode des motifs
#: fournis ; les autres valeurs l'imposent, et le calcul échoue si le motif
#: correspondant manque — un mode demandé qui se dégraderait en silence
#: produirait un total d'une autre nature sans que rien ne le signale.
FILTER_MODES = ("auto", "layer_type_filter", "geometric_type_filter", "geometric")


class EnvelopeFilterModeError(ValueError):
    """Mode de filtrage impossible à honorer avec les motifs fournis."""


def resolve_filter_mode(
    filter_mode: str | None, layer_pattern: str | None, type_pattern: str | None
) -> str:
    """Mode effectif, explicite ou déduit — et **jamais** dégradé en silence."""
    mode = (filter_mode or "auto").strip()
    if mode not in FILTER_MODES:
        raise EnvelopeFilterModeError(
            f"``filter_mode`` inconnu : {mode!r}. Valeurs acceptées : "
            + ", ".join(FILTER_MODES)
        )
    if mode == "auto":
        if layer_pattern:
            return "layer_type_filter"
        if type_pattern:
            return "geometric_type_filter"
        return "geometric"
    if mode == "layer_type_filter" and not layer_pattern:
        raise EnvelopeFilterModeError(
            "``filter_mode='layer_type_filter'`` exige ``layer_pattern`` : c'est "
            "le calque qui délimite l'enveloppe dans ce mode."
        )
    if mode == "geometric_type_filter" and not type_pattern:
        raise EnvelopeFilterModeError(
            "``filter_mode='geometric_type_filter'`` exige ``type_pattern`` : sans "
            "lui, la sélection retiendrait tous les murs extérieurs et compterait "
            "plusieurs fois la même façade sur une maquette multicouche."
        )
    if mode == "geometric" and (layer_pattern or type_pattern):
        raise EnvelopeFilterModeError(
            "``filter_mode='geometric'`` n'applique aucun motif : "
            "``layer_pattern`` / ``type_pattern`` seraient ignorés en silence. "
            "Retirer les motifs, ou choisir le mode qui les emploie."
        )
    return mode


def run(
    model,
    file_name: str,
    *,
    seuil_3f: float | None = None,
    layer_pattern: str | None = None,
    type_pattern: str | None = None,
    filter_mode: str | None = None,
) -> dict:
    """Produit le document ``envelope_quantities/v1`` (contrat bim-core).

    Trois modes de sélection des murs d'enveloppe :

    - ``layer_type_filter`` (``layer_pattern`` fourni) — sélection I3F
      **ArchiCAD** : le calque délimite l'enveloppe, ``type_pattern`` sépare
      murs extérieurs et habillages. Total façade = ``NetSideArea`` des types
      retenus, menuiseries comptées **sur ces murs**, SHAB restreinte aux pièces
      zonées hors annexes. C'est le mode qui reproduit l'extraction de référence ;
    - ``geometric_type_filter`` (``type_pattern`` seul) — maquettes **sans
      calque**, typiquement un export Revit. Les murs extérieurs sont trouvés
      géométriquement, puis ``type_pattern`` désigne la couche qui représente la
      façade : sans lui, une façade modélisée en murs superposés (structure,
      isolant, peau) serait comptée trois ou quatre fois ;
    - ``geometric`` (défaut, aucun motif) — murs marqués extérieurs, sans
      hypothèse de convention. Conservé tel quel.

    ``ratio_fac_shab`` a une définition **unique** dans les trois modes :
    ``superficie_facades_nette_m2 / shab_m2``, celle que le livrable Excel
    calcule. Elle exclut les menuiseries — deux définitions concurrentes du même
    indicateur ont déjà circulé (0,92 dans le classeur, 1,05 dans le contrat).

    Les motifs employés sont **explicites** (aucune valeur codée en dur pour un
    projet donné) et repris dans ``diagnostics.filters``, avec les types retenus
    et rejetés : la sélection est rejouable à partir du seul contrat.

    Le document est **construit directement au format V1** : il n'est jamais
    assemblé à plat puis migré.
    """
    mode = resolve_filter_mode(filter_mode, layer_pattern, type_pattern)
    if mode == "layer_type_filter":
        return _run_i3f(
            model,
            file_name,
            seuil_3f=seuil_3f,
            layer_pattern=layer_pattern,
            type_pattern=type_pattern,
        )
    if mode == "geometric_type_filter":
        return _run_geometric_type_filter(
            model,
            file_name,
            seuil_3f=seuil_3f,
            type_pattern=type_pattern,
        )

    fac = _facades(model)
    men = _menuiseries(model)
    facade_net = fac["facade_net"]
    facade_gross = round(facade_net + men["total"], 2)
    shab, n_shab = _shab(model)
    # Ratio unique : surface NETTE d'enveloppe / SHAB, menuiseries exclues.
    ratio = (facade_net / shab) if shab else None

    return {
        "schema": SCHEMA_ENVELOPE_QUANTITIES_V1,
        "source": contract_source("extract_envelope_surfaces", file_name),
        "created_at": utc_now_iso(),
        "summary": {
            # Total métier = murs extérieurs (par_type), menuiseries incluses.
            "superficie_facades_m2": facade_gross,
            "superficie_facades_nette_m2": facade_net,
            # Total « brut » avant filtrage (tous murs) — jamais le total métier.
            "superficie_calque_total_m2": fac["calque_total"],
            "superficie_menuiseries_m2": round(men["total"], 2),
            "superficie_menuiseries_fenetres_m2": round(men["fenetres"], 2),
            "superficie_menuiseries_portes_m2": round(men["portes"], 2),
            "shab_m2": round(shab, 2),
            "ratio_fac_shab": round(ratio, 3) if ratio is not None else None,
            "seuil_i3f": seuil_3f,
            "methode_facade": fac["method"],
        },
        # Décomposition métier (colonnes MOA) vs diagnostic (hors filtre).
        "par_type": fac["par_type"],
        "hors_filtre_type": fac["hors_filtre_type"],
        # Hors total métier, par construction : compteurs et détails d'appui.
        "diagnostics": {
            "filters": _filters_trace(
                mode="geometric",
                layer_pattern=None,
                type_pattern=None,
                par_type=fac["par_type"],
                hors_filtre_type=fac["hors_filtre_type"],
            ),
            "counts": {
                "n_murs_exterieurs": fac["n_ext"],
                "n_facades_fallback_geom": fac["n_geom_fallback"],
                "n_types_facade": len(fac["par_type"]),
                "n_types_hors_filtre": len(fac["hors_filtre_type"]),
                "n_menuiseries": men["n"],
                "n_pieces_shab": n_shab,
            },
            "menuiseries_detail": men["detail"],
        },
    }


def _filters_trace(
    *,
    mode: str,
    layer_pattern: str | None,
    type_pattern: str | None,
    par_type: list[dict],
    hors_filtre_type: list[dict],
) -> dict:
    """Trace du filtre appliqué, pour rejouer la sélection depuis le contrat.

    Le mode et les motifs seuls ne suffisent pas à relire un résultat : il faut
    voir **ce qu'ils ont produit**. On expose donc aussi les types retenus et
    les types rejetés — c'est ce qui permet de constater qu'une couche de façade
    manque, ou qu'un doublage a été compté, sans rouvrir l'IFC.
    """
    return {
        "mode": mode,
        "layer_pattern": layer_pattern,
        "type_pattern": type_pattern,
        "types_retenus": sorted(r["type"] for r in par_type),
        "types_rejetes": sorted(r["type"] for r in hors_filtre_type),
    }


def _run_geometric_type_filter(
    model,
    file_name: str,
    *,
    seuil_3f: float | None,
    type_pattern: str,
) -> dict:
    """Mode « murs extérieurs géométriques + filtre de type ». Voir :func:`run`."""
    type_re = re.compile(type_pattern, re.I)

    fac = _facades_by_geometric_type(model, type_re)
    # Menuiseries portées par les murs RETENUS, comme en mode I3F : compter
    # celles de toute la maquette rapporterait des baies de murs écartés.
    men = _menuiseries_of_walls(model, fac["wall_types"])
    shab, n_shab, shab_exclu = _shab_zoned(model)
    facade_net = fac["facade_net"]
    ratio = (facade_net / shab) if shab else None

    openings = men["par_type"]
    for row in fac["par_type"]:
        row["menuiseries_m2"] = round(openings.get(row["type"], 0.0), 2)

    return {
        "schema": SCHEMA_ENVELOPE_QUANTITIES_V1,
        "source": contract_source("extract_envelope_surfaces", file_name),
        "created_at": utc_now_iso(),
        "summary": {
            # Total métier = NetSideArea des types RETENUS, menuiseries exclues.
            "superficie_facades_m2": facade_net,
            "superficie_facades_nette_m2": facade_net,
            # Murs extérieurs AVANT filtre de type : sur une maquette multicouche
            # cette valeur compte plusieurs fois la même façade. Diagnostic seul.
            "superficie_calque_total_m2": fac["calque_total"],
            "superficie_menuiseries_m2": round(men["total"], 2),
            "superficie_menuiseries_fenetres_m2": round(men["fenetres"], 2),
            "superficie_menuiseries_portes_m2": round(men["portes"], 2),
            "shab_m2": round(shab, 2),
            "ratio_fac_shab": round(ratio, 4) if ratio is not None else None,
            "seuil_i3f": seuil_3f,
            "conforme_seuil": (
                bool(ratio <= seuil_3f) if (ratio is not None and seuil_3f) else None
            ),
            "methode_facade": fac["method"],
        },
        "par_type": fac["par_type"],
        "hors_filtre_type": fac["hors_filtre_type"],
        "diagnostics": {
            "filters": _filters_trace(
                mode=fac["method"],
                layer_pattern=None,
                type_pattern=type_pattern,
                par_type=fac["par_type"],
                hors_filtre_type=fac["hors_filtre_type"],
            ),
            "counts": {
                "n_murs_exterieurs": fac["n_murs_exterieurs"],
                "n_murs_retenus": fac["n_ext"],
                "n_facades_fallback_geom": fac["n_geom_fallback"],
                "n_types_facade": len(fac["par_type"]),
                "n_types_hors_filtre": len(fac["hors_filtre_type"]),
                "n_menuiseries": men["n"],
                "n_pieces_shab": n_shab,
            },
            "methode_exterieur": fac["method_exterieur"],
            "shab_types_exclus": sorted(_SHAB_EXCLUDE_I3F_TYPES)
            + [_SHAB_EXCLUDE_I3F_RAW.pattern],
            "shab_exclusions_m2": round(shab_exclu, 2),
            "menuiseries_detail": men["detail"],
        },
    }


def _run_i3f(
    model,
    file_name: str,
    *,
    seuil_3f: float | None,
    layer_pattern: str,
    type_pattern: str | None,
) -> dict:
    """Mode calque + type (extraction I3F). Voir :func:`run`."""
    layer_re = re.compile(layer_pattern, re.I)
    type_re = re.compile(type_pattern, re.I) if type_pattern else None

    fac = _facades_by_layer(model, layer_re, type_re)
    men = _menuiseries_of_walls(model, fac["wall_types"])
    shab, n_shab, shab_exclu = _shab_zoned(model)
    facade_net = fac["facade_net"]
    ratio = (facade_net / shab) if shab else None

    # Ventilation des menuiseries par type de mur porteur (colonne MOA F).
    openings = men["par_type"]
    for row in fac["par_type"]:
        row["menuiseries_m2"] = round(openings.get(row["type"], 0.0), 2)

    return {
        "schema": SCHEMA_ENVELOPE_QUANTITIES_V1,
        "source": contract_source("extract_envelope_surfaces", file_name),
        "created_at": utc_now_iso(),
        "summary": {
            # Total métier = NetSideArea des types RETENUS, menuiseries exclues
            # (elles ont leur propre total) — définition de l'extraction I3F.
            "superficie_facades_m2": facade_net,
            "superficie_facades_nette_m2": facade_net,
            # Total du calque, filtre de type inclus : retenus + hors filtre.
            "superficie_calque_total_m2": fac["calque_total"],
            "superficie_menuiseries_m2": round(men["total"], 2),
            "superficie_menuiseries_fenetres_m2": round(men["fenetres"], 2),
            "superficie_menuiseries_portes_m2": round(men["portes"], 2),
            "shab_m2": round(shab, 2),
            "ratio_fac_shab": round(ratio, 4) if ratio is not None else None,
            "seuil_i3f": seuil_3f,
            "conforme_seuil": (
                bool(ratio <= seuil_3f) if (ratio is not None and seuil_3f) else None
            ),
            "methode_facade": fac["method"],
        },
        "par_type": fac["par_type"],
        "hors_filtre_type": fac["hors_filtre_type"],
        "diagnostics": {
            # Motifs employés : la sélection est reproductible et auditable.
            "filters": _filters_trace(
                mode=fac["method"],
                layer_pattern=layer_pattern,
                type_pattern=type_pattern,
                par_type=fac["par_type"],
                hors_filtre_type=fac["hors_filtre_type"],
            ),
            "counts": {
                "n_murs_calque": fac["n_murs_calque"],
                "n_murs_sans_calque": fac["n_murs_sans_calque"],
                "n_murs_retenus": fac["n_ext"],
                "n_facades_fallback_geom": fac["n_geom_fallback"],
                "n_types_facade": len(fac["par_type"]),
                "n_types_hors_filtre": len(fac["hors_filtre_type"]),
                "n_menuiseries": men["n"],
                "n_pieces_shab": n_shab,
            },
            "shab_types_exclus": sorted(_SHAB_EXCLUDE_I3F_TYPES)
            + [_SHAB_EXCLUDE_I3F_RAW.pattern],
            "shab_exclusions_m2": round(shab_exclu, 2),
            "menuiseries_detail": men["detail"],
        },
    }


# --------------------------------------------------------------------------- #
#  Écriture du classeur .xlsx au format read_enveloppe (avp_sources)
# --------------------------------------------------------------------------- #
def write_xlsx(payload: dict, path: str) -> None:
    """**LEGACY** — classeur d'appoint, conservé pour compatibilité.

    Consomme le document ``envelope_quantities/v1`` : totaux dans ``summary``,
    détails dans ``diagnostics``.

    .. deprecated::
        Ce MCP **calcule** ; la mise en forme client (XLSX, DOCX, PDF) relève
        d'``audit-bim-i3f``, qui consomme le contrat JSON. Produire un classeur
        ici dupliquerait la charte MOA dans deux dépôts, avec deux vérités
        possibles. Remplacement : passer ``<stem>_envelope.json`` en
        ``envelope_json`` à ``generate_avp_i3f_pack`` (ou le laisser le
        résoudre seul).
    """
    warnings.warn(
        "envelope.write_xlsx est LEGACY : le flux officiel est le contrat JSON "
        "``envelope_quantities/v1``, mis en forme par audit-bim-i3f "
        "(``generate_avp_i3f_pack``).",
        DeprecationWarning,
        stacklevel=2,
    )
    import openpyxl

    summary = payload.get("summary") or {}
    diagnostics = payload.get("diagnostics") or {}
    source = payload.get("source") or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extraction surface enveloppe"

    ws["A1"] = (
        "BIMDATA — EXTRACTION SURFACE ENVELOPPE (calcul géométrique IfcOpenShell)"
    )
    ws["A2"] = f"Fichier : {source.get('ifc_file')}"

    # Bloc synthèse : libellé en A, valeur numérique en B (lu par _scan_value).
    synth = [
        ("Superficie des façades", summary.get("superficie_facades_m2")),
        ("Superficie des menuiseries", summary.get("superficie_menuiseries_m2")),
        ("SHAB", summary.get("shab_m2")),
        ("ratio FAC/SHAB", summary.get("ratio_fac_shab")),
    ]
    if summary.get("seuil_i3f") is not None:
        synth.append(("Seuil 3F 2026", summary.get("seuil_i3f")))

    ws["A4"] = "Synthèse"
    r = 5
    for label, val in synth:
        ws.cell(r, 1, label)
        ws.cell(r, 2, val)
        r += 1

    # Table détaillée des menuiseries (ancrée sur « Composant »).
    r += 1
    headers = ["Composant", "Type", "Largeur (m)", "Hauteur (m)", "Surface (m²)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(r, c, h)
    r += 1
    for d in diagnostics.get("menuiseries_detail") or []:
        ws.cell(r, 1, d.get("name"))
        ws.cell(r, 2, d.get("type"))
        ws.cell(r, 3, d.get("largeur_m"))
        ws.cell(r, 4, d.get("hauteur_m"))
        ws.cell(r, 5, d.get("surface_m2"))
        r += 1

    wb.save(path)
